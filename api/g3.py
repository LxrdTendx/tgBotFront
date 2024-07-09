import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
import win32com.client as win32
import os
from datetime import datetime
import time
start_time = time.time()

# URL вашего API
API_URL = 'http://127.0.0.1:8000'

# Получение данных фронта с id 42
front_id = 48
response = requests.get(f'{API_URL}/fronttransfers/{front_id}')
if response.status_code != 200:
    raise Exception(f'Ошибка при получении данных фронта: {response.status_code}')

front = response.json()

# Получение необходимых данных для замены
object_response = requests.get(f'{API_URL}/objects/{front["object_id"]}')
if object_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных объекта: {object_response.status_code}')
object_name = object_response.json()['name']

block_section_response = requests.get(f'{API_URL}/blocksections/{front["block_section_id"]}')
if block_section_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных блока/секции: {block_section_response.status_code}')
block_section_name = block_section_response.json()['name']

boss_response = requests.get(f'{API_URL}/users/{front["boss_id"]}')
if boss_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных босса: {boss_response.status_code}')
boss_name = boss_response.json()['full_name']

sender_response = requests.get(f'{API_URL}/users/{front["sender_id"]}')
if sender_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных получателя: {sender_response.status_code}')
sender = sender_response.json()
sender_name = sender['full_name']
organization_id_sender = sender['organization_id']

receiver_response = requests.get(f'{API_URL}/users/{front["receiver_id"]}')
if receiver_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных получателя: {receiver_response.status_code}')
receiver = receiver_response.json()
receiver_name = receiver['full_name']
organization_id_receiver = receiver['organization_id']


# Получение названия организации
organization_response1 = requests.get(f'{API_URL}/organizations/{organization_id_sender}')
if organization_response1.status_code != 200:
    raise Exception(f'Ошибка при получении данных организации: {organization_response1.status_code}')
organization_name1 = organization_response1.json()['organization']


organization_response2 = requests.get(f'{API_URL}/organizations/{organization_id_receiver}')
if organization_response2.status_code != 200:
    raise Exception(f'Ошибка при получении данных организации: {organization_response2.status_code}')
organization_name2 = organization_response2.json()['organization']

# Извлечение дня и месяца из поля approval_at
approval_at = datetime.fromisoformat(front['approval_at'])
day = approval_at.day

# Русские названия месяцев
months = ["января", "февраля", "марта", "апреля", "мая", "июня",
          "июля", "августа", "сентября", "октября", "ноября", "декабря"]
month = months[approval_at.month - 1]

# Открытие Excel-файла
excel_path = os.path.abspath('Акт_приема_передачи_фронта_работ_трехсторонний.xlsx')
workbook = load_workbook(excel_path)
worksheet = workbook.active

# Установка фиксированной ширины для первых трех колонок
worksheet.column_dimensions['A'].width = 25
worksheet.column_dimensions['B'].width = 25
worksheet.column_dimensions['C'].width = 28

# Замена плейсхолдеров в документе и установка размера шрифта
def replace_placeholder(ws, placeholder, replacement):
    font = Font(size=10)  # Создаем шрифт с размером 10
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and placeholder in cell.value:
                cell.value = cell.value.replace(placeholder, replacement)
                cell.font = font

replace_placeholder(worksheet, 'objectname', object_name)
replace_placeholder(worksheet, 'blocksectionid', block_section_name)
replace_placeholder(worksheet, 'bossname', boss_name)
replace_placeholder(worksheet, 'sendername', sender_name)
replace_placeholder(worksheet, 'receivername', receiver_name)
replace_placeholder(worksheet, 'floor', front['floor'])
replace_placeholder(worksheet, 'orgname1', organization_name1)
replace_placeholder(worksheet, 'orgname2', organization_name2)
replace_placeholder(worksheet, 'day', str(day))
replace_placeholder(worksheet, 'month', month)

# Сохранение обновленного документа в буфер
temp_excel_path = os.path.abspath('temp_updated_document.xlsx')
workbook.save(temp_excel_path)
print("Документ сохранен: ", temp_excel_path)

# Конвертация Excel в PDF
excel_app = win32.Dispatch('Excel.Application')
workbook = excel_app.Workbooks.Open(temp_excel_path)
pdf_output_path = os.path.abspath(f'{object_name}_{boss_name}_{sender_name}_Акт_приема_передачи_фронта_работ_трехсторонний.pdf')
workbook.ExportAsFixedFormat(0, pdf_output_path)
workbook.Close(False)
excel_app.Quit()

end_time = time.time()
elapsed_time = end_time - start_time

print(f'Документ успешно обновлен и конвертирован в PDF и сохранен как {pdf_output_path}')
print(f'Время выполнения: {elapsed_time:.2f} секунд')