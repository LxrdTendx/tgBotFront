import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
import win32com.client as win32
import os
from datetime import datetime

# URL вашего API
API_URL = 'http://127.0.0.1:8000'

# Получение данных фронта с id 42
front_id = 42
response = requests.get(f'{API_URL}/fronttransfers/{front_id}')
if response.status_code != 200:
    raise Exception(f'Ошибка при получении данных фронта: {response.status_code}')

front = response.json()

# Получение необходимых данных для замены
object_response = requests.get(f'{API_URL}/objects/{front["object_id"]}')
if object_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных объекта: {object_response.status_code}')
object_name = object_response.json()['name']

block_section_response = requests.get(f'{API_URL}/blocksections/{front["block_section_id"]}')
if block_section_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных блока/секции: {block_section_response.status_code}')
block_section_name = block_section_response.json()['name']

boss_response = requests.get(f'{API_URL}/users/{front["boss_id"]}')
if boss_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных босса: {boss_response.status_code}')
boss_name = boss_response.json()['full_name']

receiver_response = requests.get(f'{API_URL}/users/{front["sender_id"]}')
if receiver_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных получателя: {receiver_response.status_code}')
receiver = receiver_response.json()
receiver_name = receiver['full_name']
organization_id = receiver['organization_id']

# Получение названия организации
organization_response = requests.get(f'{API_URL}/organizations/{organization_id}')
if organization_response.status_code != 200:
    raise Exception(f'Ошибка при получении данных организации: {organization_response.status_code}')
organization_name = organization_response.json()['organization']

# Извлечение дня и месяца из поля approval_at
approval_at = datetime.fromisoformat(front['approval_at'])
day = approval_at.day

# Русские названия месяцев
months = ["января", "февраля", "марта", "апреля", "мая", "июня",
          "июля", "августа", "сентября", "октября", "ноября", "декабря"]
month = months[approval_at.month - 1]

# Открытие Excel-файла
excel_path = os.path.abspath('Акт_приема_передачи_фронта_работ_двухсторонний.xlsx')
workbook = load_workbook(excel_path)
worksheet = workbook.active

# Установка фиксированной ширины для первых трех колонок
worksheet.column_dimensions['A'].width = 25
worksheet.column_dimensions['B'].width = 25
worksheet.column_dimensions['C'].width = 28

# Замена плейсхолдеров в документе и установка размера шрифта
def replace_placeholder(ws, placeholder, replacement):
    font = Font(size=10)  # Создаем шрифт с размером 10
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and placeholder in cell.value:
                cell.value = cell.value.replace(placeholder, replacement)
                cell.font = font

replace_placeholder(worksheet, 'objectname', object_name)
replace_placeholder(worksheet, 'blocksectionid', block_section_name)
replace_placeholder(worksheet, 'bossname', boss_name)
replace_placeholder(worksheet, 'sendername', receiver_name)
replace_placeholder(worksheet, 'floor', front['floor'])
replace_placeholder(worksheet, 'orgname', organization_name)
replace_placeholder(worksheet, 'day', str(day))
replace_placeholder(worksheet, 'month', month)

# Сохранение обновленного документа в буфер
temp_excel_path = os.path.abspath('temp_updated_document.xlsx')
workbook.save(temp_excel_path)
print("Документ сохранен: ", temp_excel_path)

# Конвертация Excel в PDF
excel_app = win32.Dispatch('Excel.Application')
workbook = excel_app.Workbooks.Open(temp_excel_path)
pdf_output_path = os.path.abspath(f'{object_name}_{boss_name}_{receiver_name}_Акт_приема_передачи_фронта_работ_двухсторонний.pdf')
workbook.ExportAsFixedFormat(0, pdf_output_path)
workbook.Close(False)
excel_app.Quit()

print(f'Документ успешно обновлен и конвертирован в PDF и сохранен как {pdf_output_path}')
